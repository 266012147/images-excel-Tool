import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Font, Alignment
from PIL import Image as PILImage
import tempfile
import os
import zipfile



# 页面配置
st.set_page_config(
    page_title="图片批量导入Excel工具",
    page_icon="📷",
    layout="wide"
)



# 应用标题和说明
st.title("📷 图片批量导入Excel工具")
st.markdown("所有图片将作为浮动图放置在Excel的第二列上方")



# 侧边栏配置选项
st.sidebar.header("配置选项")

col_width = st.sidebar.number_input(
    "单元格宽度 (字符宽度)", 
    min_value=5.0, 
    max_value=100.0, 
    value=25.0, 
    step=0.1,
    help="Excel中B列的宽度"
)

row_height = st.sidebar.number_input(
    "单元格高度 (像素)", 
    min_value=20, 
    max_value=500, 
    value=100, 
    step=1,
    help="每行的高度"
)

start_row = st.sidebar.number_input(
    "起始行", 
    min_value=1, 
    max_value=1000, 
    value=2, 
    step=1,
    help="从第几行开始放置图片"
)

keep_original_size = st.sidebar.checkbox(
    "保持原始图片尺寸", 
    value=True,
    help="不缩放图片，仅调整显示大小"
)



# 文件上传区域
st.subheader("上传图片")
uploaded_files = st.file_uploader(
    "选择图片文件",
    type=['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp'],
    accept_multiple_files=True,
    help="支持 JPG, PNG, GIF, BMP, WEBP 格式"
)



# 显示已上传的文件信息
if uploaded_files:
    st.success(f"已选择 {len(uploaded_files)} 张图片")
    
    # 显示文件列表
    file_info = []
    total_size = 0
    
    for i, file in enumerate(uploaded_files):
        file_size = len(file.getvalue()) / 1024  # KB
        total_size += file_size
        file_info.append({
            "文件名": file.name,
            "大小 (KB)": f"{file_size:.1f}",
            "类型": file.type
        })
    
    # 显示文件表格
    if file_info:
        df_files = pd.DataFrame(file_info)
        st.dataframe(df_files, use_container_width=True)
        st.info(f"总大小: {total_size:.1f} KB")



# 生成Excel按钮
if uploaded_files:
    if st.button("🚀 生成Excel文件", type="primary", use_container_width=True):
        try:
            # 创建进度条
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "图片列表"
            
            # 设置第二列宽度 (B列)
            ws.column_dimensions['B'].width = col_width
            
            # 添加标题行
            ws['A1'] = '序号'
            ws['B1'] = '图片'
            ws['A1'].font = Font(bold=True, size=12)
            ws['B1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 创建临时目录
            with tempfile.TemporaryDirectory() as temp_dir:
                # 处理每张图片
                for idx, uploaded_file in enumerate(uploaded_files):
                    progress = (idx + 1) / len(uploaded_files)
                    progress_bar.progress(progress)
                    status_text.text(f"正在处理第 {idx + 1}/{len(uploaded_files)} 张图片: {uploaded_file.name}")
                    
                    row_num = start_row + idx
                    
                    # 设置行高（像素转磅：1像素 ≈ 0.75磅）
                    ws.row_dimensions[row_num].height = row_height * 0.75
                    
                    # 写入序号
                    cell_a = ws.cell(row=row_num, column=1, value=idx + 1)
                    cell_a.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 读取图片数据
                    file_bytes = uploaded_file.getvalue()
                    
                    # 使用PIL打开图片获取信息
                    pil_img = PILImage.open(io.BytesIO(file_bytes))
                    original_width, original_height = pil_img.size
                    
                    # 计算单元格的实际像素尺寸
                    cell_width_px = col_width * 7  # Excel列宽单位转像素
                    cell_height_px = row_height
                    
                    # 保存临时文件
                    original_ext = os.path.splitext(uploaded_file.name)[1].lower()
                    if original_ext not in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
                        original_ext = '.png'
                    
                    temp_original_path = os.path.join(temp_dir, f'original_{idx}{original_ext}')
                    with open(temp_original_path, 'wb') as f:
                        f.write(file_bytes)
                    
                    if keep_original_size:
                        # 保持原始图片，不进行缩放处理
                        xl_img = XLImage(temp_original_path)
                        
                        # 计算显示尺寸，保持比例适应单元格
                        img_ratio = original_width / original_height
                        cell_ratio = cell_width_px / cell_height_px
                        
                        if img_ratio > cell_ratio:
                            # 图片更宽，以宽度为准
                            display_width = int(cell_width_px * 0.9)
                            display_height = int(display_width / img_ratio)
                        else:
                            # 图片更高，以高度为准
                            display_height = int(cell_height_px * 0.9)
                            display_width = int(display_height * img_ratio)
                        
                        # 设置图片在Excel中的显示大小
                        xl_img.width = display_width
                        xl_img.height = display_height
                        
                    else:
                        # 需要缩放图片时，使用高质量缩放
                        img_ratio = original_width / original_height
                        cell_ratio = cell_width_px / cell_height_px
                        
                        if img_ratio > cell_ratio:
                            new_width = int(cell_width_px * 0.9)
                            new_height = int(new_width / img_ratio)
                        else:
                            new_height = int(cell_height_px * 0.9)
                            new_width = int(new_height * img_ratio)
                        
                        # 确保不超出单元格
                        new_width = min(new_width, int(cell_width_px * 0.95))
                        new_height = min(new_height, int(cell_height_px * 0.95))
                        
                        # 使用高质量LANCZOS算法缩放
                        pil_img_resized = pil_img.resize((new_width, new_height), PILImage.LANCZOS)
                        
                        # 处理透明图片
                        if pil_img_resized.mode in ('RGBA', 'LA', 'P'):
                            background = PILImage.new('RGB', pil_img_resized.size, (255, 255, 255))
                            if pil_img_resized.mode == 'P':
                                pil_img_resized = pil_img_resized.convert('RGBA')
                            if pil_img_resized.mode == 'RGBA':
                                background.paste(pil_img_resized, mask=pil_img_resized.split()[3])
                            else:
                                background.paste(pil_img_resized)
                            pil_img_resized = background
                        
                        # 高质量保存为PNG
                        temp_resized_path = os.path.join(temp_dir, f'resized_{idx}.png')
                        pil_img_resized.save(temp_resized_path, format='PNG', optimize=False)
                        
                        # 使用缩放后的文件
                        xl_img = XLImage(temp_resized_path)
                        xl_img.width = new_width
                        xl_img.height = new_height
                    
                    # 创建浮动图定位（使用OneCellAnchor）
                    # 定位到B列对应行的单元格
                    col = 2  # B列
                    row = row_num
                    
                    # 将像素转换为EMU（Excel使用的单位）
                    img_width_emu = pixels_to_EMU(xl_img.width)
                    img_height_emu = pixels_to_EMU(xl_img.height)
                    
                    # 计算单元格尺寸（EMU）
                    cell_width_emu = pixels_to_EMU(cell_width_px)
                    cell_height_emu = pixels_to_EMU(cell_height_px)
                    
                    # 计算居中偏移量（让图片在单元格内居中显示）
                    offset_x = max(0, (cell_width_emu - img_width_emu) // 2)
                    offset_y = max(0, (cell_height_emu - img_height_emu) // 2)
                    
                    # 创建锚点定位
                    marker = AnchorMarker(col=col-1, colOff=offset_x, row=row-1, rowOff=offset_y)
                    
                    # 创建正确的尺寸对象
                    ext = XDRPositiveSize2D(img_width_emu, img_height_emu)
                    
                    # 创建锚点
                    anchor = OneCellAnchor(_from=marker, ext=ext)
                    
                    # 将锚点设置给图片
                    xl_img.anchor = anchor
                    
                    # 添加图片到工作表
                    ws.add_image(xl_img)
                    
                    # 设置B列单元格对齐方式（保持单元格样式）
                    cell_b = ws.cell(row=row_num, column=2, value='')
                    cell_b.alignment = Alignment(horizontal='center', vertical='center')
                
                # 调整第一列宽度
                ws.column_dimensions['A'].width = 10
                
                # 保存到内存
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # 更新进度条
                progress_bar.progress(1.0)
                status_text.text("✅ Excel文件生成完成！")
                
                # 提供下载
                st.success("Excel文件生成成功！图片以浮动图形式显示。")
                
                st.download_button(
                    label="📥 下载Excel文件",
                    data=output,
                    file_name="images_in_excel.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
        except Exception as e:
            st.error(f"生成Excel时出错: {str(e)}")
            st.stop()

else:
    st.info("👆 请先上传图片文件")



# 使用说明
with st.expander("📖 使用说明"):
    st.markdown("""
    ### 功能说明
    - **批量处理**: 支持一次性上传多张图片
    - **浮动图片**: 图片作为浮动对象放置在单元格上方，可自由拖动
    - **保持画质**: 可选择保持原始图片尺寸，避免压缩损失
    - **智能适配**: 自动计算图片比例，适配单元格大小
    
    ### 操作步骤
    1. 在左侧配置单元格大小和起始行
    2. 上传图片文件（支持拖拽）
    3. 点击"生成Excel文件"按钮
    4. 下载生成的Excel文件
    
    ### 支持格式
    - JPG、JPEG、PNG、GIF、BMP、WEBP
    
    ### 浮动图特点
    - 图片浮在单元格上方，不嵌入单元格内
    - 可以自由拖动和调整大小
    - 保持原始清晰度
    """)



# 依赖说明
with st.expander("🔧 技术信息"):
    st.markdown("""
    ### 依赖库
    ```python
    streamlit
    openpyxl
    Pillow
    pandas
    ```
    
    ### 运行方式
    ```bash
    streamlit run app.py
    ```
    """)
